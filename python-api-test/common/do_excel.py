from openpyxl import Workbook
from openpyxl import load_workbook
# s=Workbook() #新建一个文件
# s.save(r'D:\pycharm code\python-api-test\datas\test.xlsx') #保存文件，否则无法创建
class Case:
    def __init__(self):
        self.case_id=  None
        self.title = None
        self.url = None
        self.datas = None
        self.method = None
        self.expected = None
        self.response = None
        self.test_report = None

class DoExcel:
    def __init__(self,filename):
        self.filename=filename
        # self.sheetname=sheetname

    def read(self,sheetname):
        file=load_workbook(self.filename)
        sheet = file[sheetname]
        list_data=[]
        for i in range(2,sheet.max_row+1):
            case_data=Case()
            case_data.case_id=sheet.cell(i,1).value
            case_data.title=sheet.cell(i,2).value
            case_data.url=sheet.cell(i,3).value
            case_data.datas=sheet.cell(i,4).value
            case_data.method=sheet.cell(i,5).value
            case_data.expected=sheet.cell(i,6).value
            list_data.append(case_data)
        return list_data
    def input_data(self,row,column,value,sheetname):
        file = load_workbook(self.filename)
        sheet = file[sheetname]
        sheet.cell(row,column,value=value)
        file.save(self.filename)

if __name__ == '__main__':
    s=DoExcel(r'..\datas\test.xlsx')
    # q=s.read()
    q=s.input_data(2,7,'接口自动化写入ceshi','Sheet')
    # print(q)




